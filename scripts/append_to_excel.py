import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

CSV_PATH = Path("data/daily_output_log.csv")
XLSX_PATH = Path("data/September Averages.xlsx")
SHEET_NAME = "Daily by Shifts"
SUMMARY_FILE = Path("sync_summary.md")
UNIQUE_KEY = ["Machine Name", "Date", "Shift"]

def append_or_update_rows():
    csv_df = pd.read_csv(CSV_PATH)
    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    # Load existing Excel data
    existing_df = pd.DataFrame(ws.values)
    header = existing_df.iloc[0]
    existing_df = pd.DataFrame(existing_df.values[1:], columns=header)

    # Normalize types
    for col in UNIQUE_KEY + ["Notes", "No Schedule"]:
        if col in csv_df.columns:
            csv_df[col] = csv_df[col].astype(str)
        if col in existing_df.columns:
            existing_df[col] = existing_df[col].astype(str)

    # Merge on key, suffix old columns
    merged = existing_df.merge(csv_df, on=UNIQUE_KEY, how="outer", suffixes=("_old", ""))
    update_count = 0
    new_count = 0

    for col in csv_df.columns:
        if col not in UNIQUE_KEY:
            merged[col] = merged[col].combine_first(merged[f"{col}_old"])
            merged.drop(columns=[f"{col}_old"], inplace=True, errors="ignore")

    # Count updates/new before dedupe
    before = len(existing_df)
    after = len(merged)
    new_count = max(0, after - before)

    # Drop duplicates keeping latest
    merged = merged.drop_duplicates(subset=UNIQUE_KEY, keep="last")

    # Determine updates by comparing overlapping keys
    overlap_keys = existing_df.merge(csv_df, on=UNIQUE_KEY, how="inner")
    update_count = len(overlap_keys)

    merged = merged.sort_values(by=["Date", "Shift", "Machine Name"])

    # Write back to Excel
    wb.remove(ws)
    ws_new = wb.create_sheet(SHEET_NAME)
    ws_new.append(list(merged.columns))
    for _, row in merged.iterrows():
        ws_new.append(list(row))
    wb.save(XLSX_PATH)

    # Write markdown summary
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    with open(SUMMARY_FILE, "w") as f:
        f.write("## ✅ Nightly Sync Summary\n")
        f.write(f"**Timestamp:** {timestamp}\n\n")
        f.write(f"- {update_count} rows updated\n")
        f.write(f"- {new_count} new rows appended\n")
        f.write(f"- {len(csv_df)} total rows processed\n")

    # Exit codes for workflow logic
    if new_count == 0 and update_count == 0:
        print("No new or updated rows detected.")
        return False
    else:
        print(f"✅ {new_count} new rows appended, {update_count} updated.")
        return True

if __name__ == "__main__":
    changed = append_or_update_rows()
    import sys
    sys.exit(0)
